#!/usr/bin/env python3
# -*- coding: utf-8 -*-

# SPDX-License-Identifier: BSD-3-Clause
# gnpy.tools.worker_utils: Utilities for reading and writing XLS, XLSX
# Copyright (C) 2025 Telecom Infra Project and GNPy contributors
# see AUTHORS.rst for a list of contributors

"""
gnpy.tools.xls_utils
====================

This module contains utilities for reading and writing XLS, XLSX

"""
from pathlib import Path
from typing import Generator, Tuple, List, Union, Optional, Iterator, Callable
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell as OpenpyxlCell
from openpyxl.cell.read_only import ReadOnlyCell as OpenpyxlReadOnlyCell
from openpyxl.utils.exceptions import InvalidFileException
from xlrd import Book, open_workbook, XL_CELL_EMPTY
from xlrd.sheet import Sheet as XlrdSheet, Cell as XlrdCell
from xlrd.biffh import XLRDError

SheetType = Union[Worksheet, XlrdSheet]
WorkbookType = Union[Workbook, Book]
CellType = Union[OpenpyxlCell, OpenpyxlReadOnlyCell, XlrdCell]
XLS_EXCEPTIONS = (InvalidFileException, KeyError, XLRDError)


def generic_open_workbook(file_path: Union[str, Path]) -> Tuple[WorkbookType, bool]:
    """Open an Excel file supporting both XLS or XLSX.

    :param file_path: Path of excel file
    :type file_path: Union[str, Path]
    :return: Tuple (workbook, is_xlsx) where is_xlsx inidcate if the file is XLSX or not
    :rtype: Tuple[WorkbookType, bool]
    """
    file_path = Path(file_path) if isinstance(file_path, str) else file_path

    if file_path.suffix.lower() in ['.xlsx', '.xlsm']:
        return load_workbook(file_path, read_only=True, data_only=True), True
    return open_workbook(file_path), False


def get_sheet(workbook: WorkbookType,
              sheet_name: str,
              is_xlsx: bool) -> SheetType:
    """Get the Excel Sheet by name

    :param workbook: Opened Excel workbook
    :type workbook: WorkbookType
    :param sheet_name: Sheet name
    :type sheet_name: SheetType
    :param is_xlsx: True if this is an XLSX workbook, False if XLS
    :type is_xlsx: bool
    :return: Excel sheet
    :rtype: SheetType
    """
    if is_xlsx:
        return workbook[sheet_name]
    return workbook.sheet_by_name(sheet_name)


def get_cell_value(sheet: SheetType, row: int, col: int, is_xlsx: bool) -> Optional[Union[str, int, float]]:
    """Get the cell value

    :param sheet: Excel sheet
    :type sheet: SheetType
    :param row: Line index (0-based)
    :type row: int
    :param col: Column index (0-based)
    :type: int
    :param is_xlsx: True if this is an XLSX workbook, False if XLS
    :type is_xlsx: bool
    :return: cell value
    :rtype: Optional[Union[str, int, float]]
    """
    if is_xlsx:
        # openpyxl uses a 1-based index
        cell = sheet.cell(row=row + 1, column=col + 1)
        return cell.value
    # xlrd uses a 0-base index
    return sheet.cell(row, col).value


def get_row(sheet: SheetType, row_index: int, is_xlsx: bool, get_rows=None) -> List[CellType]:
    """Get row in a workbook sheet.

    :param sheet: Excel sheet
    :type sheet: SheetType
    :param row_index: Line index (0-based)
    :type row_index: int
    :param is_xlsx: True si c'est un fichier XLSX, False si XLS
    :param is_xlsx: True if this is an XLSX workbook, False if XLS
    :type is_xlsx: bool
    :param get_rows: Optional function that returns preloaded rows (from fast_get_sheet_rows)
    :type get_rows: Optional[Callable]
    :return: List row cells
    :rtype: List[CellType]
    """
    if is_xlsx:
        if get_rows is not None:
            # use fast access with aclosure function
            rows = get_rows()
        else:
            rows = list(sheet.rows)
        return rows[row_index] if row_index < len(rows) else []
    return sheet.row(row_index)


def fast_get_sheet_rows(sheet: Worksheet) -> Callable:
    """Preloads all rows from an Excel sheet for fast access.

    This function loads the sheet data only once and returns a function
    that provides access to this preloaded data without having to query
    the Excel sheet on each access, which significantly improves performance,
    particularly with openpyxl.

    :param sheet: Excel worksheet (openpyxl.worksheet.worksheet.Worksheet object)
    :type sheet: Worksheet
    :return: Function that returns the preloaded rows
    :rtype: Callable[[], List[Tuple[Cell, ...]]]

    Usage example:
        > get_rows = fast_get_sheet_rows(sheet)
        > rows = get_rows()  # Access to preloaded data
        > first_row = rows[0]  # First row
    """
    # Load all sheet rows into memory only once
    # This operation can be expensive, but it's performed only once
    # load the rows only once.
    preloaded_data = list(sheet.rows)

    def get_rows():
        """Inner function (clodure function) that returns the preloaded data.

        This function doesn't reload the data on each call,
        it simply returns the reference to the already loaded data.

        :return: List of preloaded rows
        :rtype: List[Tuple[Cell, ...]]
        """
        return preloaded_data
    return get_rows


def get_row_slice(sheet: SheetType, row_index: int, start_col: int, end_col: int, is_xlsx: bool,
                  get_rows: Callable = None) -> Union[Tuple[CellType], List[CellType]]:
    """Get a row slice.

    :param sheet: Excel sheet
    :type sheet: SheetType
    :param row_index: Line index (0-based)
    :type row_index: int
    :param start_col: Index of start column (0-based)
    :type start_col: int
    :param end_col: Index of end column (0-based)
    :type end_col: int
    :param is_xlsx: True if this is an XLSX workbook, False if XLS
    :type is_xlsx: bool
    :param get_rows: Optional function that returns preloaded rows (from fast_get_sheet_rows)
    :type get_rows: Optional[Callable]
    :return: List of cells in the selected slice
    :rtype: List[CellType]
    """
    if is_xlsx:
        if get_rows is not None:
            rows = get_rows()
        else:
            rows = list(sheet.rows)
        return rows[row_index][start_col:end_col] if row_index < len(rows) else []
    return sheet.row_slice(row_index, start_col, end_col)


def convert_empty(cell_value: Optional[Union[str, int, float]]) -> Optional[Union[str, int, float]]:
    """Convert empty string into None

    :param cell_value: Cell value
    :type cell_value: Optional[Union[str, int, float]]

    >>> convert_empty('')

    >>> convert_empty('data')
    'data'

    >>> convert_empty(123)
    123
    """
    if cell_value == '':
        return None
    return cell_value


def get_num_rows(sheet: SheetType, is_xlsx: bool, get_rows: Callable = None) -> int:
    """Get the number of lines of an Excel sheet. Note that openpyxl in read_only mode can return "ghost" rows
    at the end (ReadOnlyCell cells with no actual value but formatting information even for empty rows).

    :param sheet: Excel sheet
    :type sheet: SheetType
    :param is_xlsx: True if this is an XLSX workbook, False if XLS
    :type is_xlsx: bool
    :param get_rows: Optional function that returns preloaded rows (from fast_get_sheet_rows)
    :type get_rows: Optional[Callable]
    :return: Number of lines
    :rtype: int
    """
    if is_xlsx:
        if get_rows is not None:
            return len(list(get_rows()))
        else:
            return len(list(sheet.rows))
    return sheet.nrows


def is_type_cell_empty(cell, is_xlsx: bool) -> bool:
    """Check is a cell is empty.

    :param sheet: Excel sheet
    :type sheet: SheetType
    :param row: Line index (0-based)
    :type row: int
    :param col: Column index (0-based)
    :type: int
    :param is_xlsx: True if this is an XLSX workbook, False if XLS
    :type is_xlsx: bool
    :return: True if cell is empty, else returns False
    :rtype: bool
    """
    if is_xlsx:
        return cell.value in [None, '']
    return cell.ctype == XL_CELL_EMPTY


def get_sheet_name(sheet: SheetType, is_xlsx: bool) -> str:
    """Get the name of the current sheet

    :param sheet: Excel sheet
    :type sheet: SheetType
    :param is_xlsx: True if this is an XLSX workbook, False if XLS
    :type is_xlsx: bool
    :return: Name of the sheet
    :rtype: str
    """
    if is_xlsx:
        return sheet.title
    return sheet.name


def all_rows(sh: Worksheet, is_xlsx: bool, start: int = 0, get_rows: Callable = None) -> Generator[list, None, None]:
    """Returns all rows of the xls(x) sheet starting from start row.

    :param sh: sheet: Excel sheet
    :type sheet: SheetType
    :param start: The starting row index (0-based).
    :type start: int
    :param get_rows: Optional function that returns preloaded rows (from fast_get_sheet_rows)
    :type get_rows: Optional[Callable]
    :return: A generator yielding all rows from the specified starting index.
    :rtype: Generator[list, None, None]
    """
    return (get_row(sh, x, is_xlsx, get_rows) for x in range(start, get_num_rows(sh, is_xlsx, get_rows)))


def correct_cell_int_to_str(v: Optional[Union[str, int, float]]) -> Optional[Union[str, int, float]]:
    """Ensure that int values in "id" cells are read as strings containing the int and
    do not use the automatic float conversion from xlrd or openpyxl

    :param v: cell value to convert
    :type v: Optional[Union[str, int, float]]
    :return: corrected cell value
    :rtype: Optional[Union[str, int, float]]

    >>> correct_cell_int_to_str(123)
    '123'
    >>> correct_cell_int_to_str(123.0)
    '123'
    >>> correct_cell_int_to_str('abc')
    'abc'
    >>> correct_cell_int_to_str(None)

    """
    if not isinstance(v, str) and v is not None:
        value = str(int(v))
        if value.endswith('.0'):
            value = value[:-2]
    else:
        value = v
    return value


def get_all_sheets(workbook: WorkbookType, is_xlsx: bool) -> Iterator[SheetType]:
    """Get all sheets from an Excel workbook.

    :param workbook: Opened Excel workbook
    :type workbook: WorkbookType
    :param is_xlsx: True if this is an XLSX workbook, False if XLS
    :type is_xlsx: bool
    :return: Iterator of all sheets in the workbook
    :rtype: Iterator[SheetType]
    """
    if is_xlsx:
        for sheet in workbook.worksheets:
            yield sheet
    else:
        for i in range(workbook.nsheets):
            yield workbook.sheet_by_index(i)


def get_sheet_names(workbook: WorkbookType, is_xlsx: bool) -> List[str]:
    """Get all sheet names from an Excel workbook.

    :param workbook: Opened Excel workbook
    :type workbook: WorkbookType
    :param is_xlsx: True if this is an XLSX workbook, False if XLS
    :type is_xlsx: bool
    :return: List of sheet names
    :rtype: List[str]
    """
    if is_xlsx:
        return workbook.sheetnames
    return workbook.sheet_names()
